"""
Data formatting module for GabeDA exports.

Single Responsibility: Format data for export ONLY
- Adjust column widths
- Apply Excel formatting (autofilter, etc.)
- Does NOT export data or interact with context
"""

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelFormatter:
    """
    Formats Excel workbooks.

    Responsibilities:
    - Enable autofilter on sheets
    - Auto-adjust column widths
    - Apply Excel formatting

    Does NOT:
    - Export data (ExcelExporter does this)
    - Interact with context
    - Load data
    """

    def __init__(self, max_width: int = 50, min_width: int = 10):
        """
        Initialize formatter with width constraints.

        Args:
            max_width: Maximum column width in characters (default: 50)
            min_width: Minimum column width in characters (default: 10)
        """
        self.max_width = max_width
        self.min_width = min_width

    def format_workbook(self, file_path: str) -> None:
        """
        Apply formatting to all sheets in workbook.

        Args:
            file_path: Path to Excel file

        Side effects:
            - Modifies Excel file in place
            - Adds autofilter to all sheets
            - Adjusts column widths
        """
        logger.info(f"Formatting workbook: {file_path}")

        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._format_sheet(ws)
            logger.debug(f"Formatted sheet: {sheet_name}")

        wb.save(file_path)
        wb.close()

        logger.info(f"Workbook formatting complete: {file_path}")

    def _format_sheet(self, worksheet: Worksheet) -> None:
        """
        Apply formatting to a single worksheet.

        Args:
            worksheet: openpyxl worksheet object

        Side effects:
            - Adds autofilter if data exists
            - Adjusts column widths
        """
        # Only format if sheet has data
        if worksheet.max_row > 0 and worksheet.max_column > 0:
            # Add autofilter
            worksheet.auto_filter.ref = worksheet.dimensions
            logger.debug(f"Added autofilter to {worksheet.title}")

            # Adjust column widths
            self._adjust_column_widths(worksheet)

    def _adjust_column_widths(self, worksheet: Worksheet) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            worksheet: openpyxl worksheet object

        Side effects:
            - Modifies worksheet column dimensions
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            # Find max length in this column
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            # Set adjusted width (add extra space)
            adjusted_width = min(
                max(max_length + 2, self.min_width),
                self.max_width
            )
            worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.debug(f"Adjusted column widths for {worksheet.title}")
